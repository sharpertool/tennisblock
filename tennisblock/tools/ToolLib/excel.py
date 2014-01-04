#!/usr/bin/env python

import os
import xlrd

from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter

from openpyxl.cell import get_column_letter
import openpyxl

class ExcelRow(object):

    def __init__(self,hdgs,rowData):
        for hdg in hdgs:
            self.__dict__[hdg] = rowData[hdg]

class ExcelWorksheet(object):

    def __init__(self,name):

        self.name = name
        self.headings = []
        self.rows = []

    def setHeadings(self,cols):

        self.headings = cols

    def insertRow(self,rowData):

        self.rows.append(ExcelRow(self.headings,rowData))

    def __getitem__(self,key):

        # Negative indexes from the end.
        if key < 0:
            key += len(self.rows)

        if key >= len(self.rows):
            raise IndexError, "The index (%d) is out of range." % key
        return self.rows[key]


class Excel(object):

    def __init__(self):

        self.isxlsx = False
        self.wb = None
        self.theFile = None
        self.sheets = []
        self.requiredColumns = []

    def setRequiredColumns(self,cols):
        """
        Set the list of required columns to determine if a row is
        valid.
        """
        self.requiredColumns = cols

    def getSheet(self,wb,name):

        if self.isxlsx:
            ws = wb.get_sheet_by_name(name)
        else:
            ws = wb.sheet_by_name(name)

        return ws

    def getValues(self,ws,row,cols=None):

        if self.isxlsx:
            wsrow = ws.rows[row]
            values = []
            for cell in wsrow:
                val = cell.value
                if not val:
                    val = ''
                values.append(val)
            return values
        else:
            return ws.row_values(row)

    def getRowCount(self,ws):

        if self.isxlsx:
            return ws.get_highest_row()
        else:
            return ws.nrows

    def openWorkbook(self,xlfile):
        """
        """

        self.theFile = os.path.expanduser(xlfile)
        self.wb = None

        openpyxl.style.NumberFormat.FORMAT_DATE_TIME6 = "h:mm:ss.s"

        try:
            (root,ext) = os.path.splitext(self.theFile)
            if ext == '.xlsx' or ext == '.xlsm':
                self.isxlsx = True
                self.wb = openpyxl.load_workbook(self.theFile)
                self.sheets = self.wb.get_sheet_names()
            else:
                self.isxlsx = False
                self.wb = xlrd.open_workbook(self.theFile)
                #self.sheets = self.wb.get_sheet_names()

        except Exception as e:
            print("Excel Read module could not open the file %s" % self.theFile)
            return self.wb

        return self.wb


    def openSheet(self):
        """
        """

        for sheet in self.sheets:
            print "Sheet %s" % sheet

            ws = self.getSheet(self.wb,sheet)
            if ws == None:
                print("Worksheet %s not found in %s" % (sheet,self.theFile))

            else:
                self.processSheet(ws)

    def importExcelFile(self,xlfile,sheets):
        """
        Generic import function.

        xlfile is the name of the xlfile, preferably .xlsx format.

        'sheets' - list of sheets to import.

        For each sheet, we read the used column headings, then import all used rows
        into a data structure for that worksheet.


        """

        self.xlfile = xlfile
        print("Importing Excel Data file %s" % xlfile)

        self.openWorkbook(xlfile)

        wbData = {}

        for sheet in sheets:
            sheetData = self.importExcelWorksheet(sheet)
            wbData[sheet['field']] = sheetData

        return wbData

    def importExcelWorksheet(self,sheet):

        wsData = ExcelWorksheet(sheet['name'])
        if sheet.has_key('requiredCols'):
            self.setRequiredColumns(sheet['requiredCols'])
        else:
            self.setRequiredColumns([])

        ws = self.getSheet(self.wb,sheet['name'])
        if ws:
            rowHeadings = self.getValues(ws,0)

            wsData.setHeadings(rowHeadings)

            nrows = self.getRowCount(ws)
            for rowid in range(1,nrows):
                theRow = self.getValues(ws,rowid)
                rowData = dict(zip(rowHeadings, theRow))
                if self.isValidRow(rowData):
                    wsData.insertRow(rowData)

        return wsData


    def isValidRow(self,rowData):
        """
        Basic function to determine if an excel row is valid.
        This function just make sure a list of required columns
        is not blank. For more complex validity, override this
        function.
        """

        valid = True
        for col in self.requiredColumns:
            if not rowData.has_key(col):
                valid = False
            else:
                if rowData[col] == "":
                    valid = False

        return valid



def main():
    """
    Test Drive.

    """
    pass


if __name__ == '__main__':
    main()
